"""Web crawling and scraping tools."""

from __future__ import annotations

import io
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from langchain_core.tools import tool

HEADERS = {
    "User-Agent": "Vision1M-Scorecard-Bot/1.0 (research; University of Waterloo)"
}

MIN_CONTENT_LENGTH = 200

JS_REQUIRED_PHRASES = [
    "needs javascript",
    "enable javascript",
    "requires javascript",
    "javascript is required",
    "javascript must be enabled",
    "please enable javascript",
    "this site requires javascript",
    "you need to enable javascript",
    "noscript",
]

JS_FRAMEWORK_MARKERS = [
    "__NEXT_DATA__", "__NUXT__", "react", "angular", "vue", "ember",
    "window.__INITIAL_STATE__", "window.__DATA__",
]

DATA_EXTENSIONS = {
    ".json", ".csv", ".xlsx", ".xls", ".xml",
    ".pdf", ".zip", ".geojson", ".tsv", ".parquet",
}


# ── Internal helpers ─────────────────────────────────────────────────────────

def _looks_like_js_rendered(html: str) -> bool:
    """Heuristic: page likely needs JS rendering."""
    html_lower = html.lower()

    for phrase in JS_REQUIRED_PHRASES:
        if phrase in html_lower:
            return True

    soup = BeautifulSoup(html, "html.parser")

    noscript_tags = soup.find_all("noscript")
    for tag in noscript_tags:
        if len(tag.get_text(strip=True)) > 20:
            return True

    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
        tag.decompose()
    text = soup.get_text(separator=" ", strip=True)

    if len(text) < MIN_CONTENT_LENGTH:
        return True

    body = soup.find("body")
    if body and len(body.find_all(True)) < 5:
        return True

    for marker in JS_FRAMEWORK_MARKERS:
        if marker in html_lower and len(text) < 500:
            return True

    return False


def _fetch_with_playwright(url: str) -> str:
    """Fetch a page using Playwright for JS-rendered content."""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        try:
            page.goto(url, timeout=60_000, wait_until="networkidle")
            for selector in ["table", "[data-react-class]", ".data-table", "#app", "[role='grid']"]:
                try:
                    page.wait_for_selector(selector, timeout=5_000)
                    break
                except Exception:
                    continue
            page.wait_for_timeout(3_000)
            content = page.content()
        finally:
            browser.close()
    return content


def _get_html(url: str) -> str:
    """Fetch HTML, using Playwright if JS rendering is needed."""
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    if _looks_like_js_rendered(resp.text):
        try:
            return _fetch_with_playwright(url)
        except Exception as exc:
            return f"<!--PLAYWRIGHT_ERROR:{exc}-->\n{resp.text}"

    return resp.text


def _extract_text_from_html(html: str) -> str:
    """Strip tags and return clean text."""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
        tag.decompose()
    return soup.get_text(separator="\n", strip=True)


def _extract_tables_from_html(html: str) -> str:
    """Extract HTML tables as pipe-separated rows."""
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return ""

    parts = []
    for i, table in enumerate(tables):
        rows = []
        for tr in table.find_all("tr"):
            cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"=== Table {i+1} ===\n" + "\n".join(rows))

    return "\n\n".join(parts)


def _extract_data_links(html: str, base_url: str) -> str:
    """Extract links to downloadable data files from HTML."""
    soup = BeautifulSoup(html, "html.parser")

    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        full_url = urljoin(base_url, href)
        label = a.get_text(strip=True)

        href_lower = href.lower().split("?")[0]
        if any(href_lower.endswith(ext) for ext in DATA_EXTENSIONS):
            links.append({"url": full_url, "label": label or href_lower.split("/")[-1]})
            continue

        label_lower = label.lower()
        if any(kw in label_lower for kw in ["download", "dataset", "data set", "csv", "json", "excel", "spreadsheet"]):
            links.append({"url": full_url, "label": label})

    if not links:
        return ""

    seen = set()
    unique = []
    for link in links:
        if link["url"] not in seen:
            seen.add(link["url"])
            unique.append(link)

    parts = [f"- [{link['label']}]({link['url']})" for link in unique]
    return f"Data links found ({len(unique)}):\n" + "\n".join(parts)


def _extract_pdf(url: str) -> str:
    """Download and extract text from a PDF."""
    import fitz  # pymupdf

    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    doc = fitz.open(stream=resp.content, filetype="pdf")
    parts = []
    for page in doc:
        parts.append(page.get_text())
    doc.close()

    text = "\n".join(parts)
    return text if text.strip() else "PDF contains no extractable text (may be scanned/image-based)."


def _extract_xlsx(url: str) -> str:
    """Download and extract text from an Excel file."""
    import openpyxl

    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 50:
                rows.append("... truncated at 50 rows")
                break
            row_str = " | ".join(str(c) if c is not None else "" for c in row)
            if row_str.strip(" |"):
                rows.append(row_str)
        if rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
    wb.close()

    return "\n\n".join(parts) if parts else "No data found in spreadsheet."


# ── Tools ────────────────────────────────────────────────────────────────────

def _classify_http_error(exc: httpx.HTTPStatusError) -> tuple[str, str]:
    """Map an HTTP status code to an error_code and message."""
    status = exc.response.status_code
    if status == 401:
        return "auth_required", f"Authentication required (HTTP {status})"
    if status == 403:
        return "forbidden", f"Access forbidden (HTTP {status})"
    if status == 429:
        return "rate_limited", f"Rate limited (HTTP {status})"
    if 400 <= status < 500:
        return "4xx", f"Client error (HTTP {status})"
    if 500 <= status < 600:
        return "5xx", f"Server error (HTTP {status})"
    return "4xx", f"HTTP error (status {status})"


def _classify_connection_error(exc: Exception) -> tuple[str, str]:
    """Map a connection-level exception to an error_code and message."""
    msg = str(exc).lower()
    if "timeout" in msg or "timed out" in msg:
        return "timeout", f"Request timed out: {exc}"
    if "dns" in msg or "name or service not known" in msg or "nodename" in msg:
        return "dns", f"DNS resolution failed: {exc}"
    if "ssl" in msg or "tls" in msg or "certificate" in msg:
        return "ssl", f"SSL/TLS error: {exc}"
    if "redirect" in msg:
        return "redirect_loop", f"Redirect error: {exc}"
    if "reset" in msg or "connection" in msg:
        return "connection_reset", f"Connection error: {exc}"
    return "connection_reset", f"Network error: {exc}"


def _log_error(url: str, error_code: str, error_message: str, http_status: int | None = None, preview: str = ""):
    """Best-effort log to Mongo. Never raises."""
    try:
        from storage.source_store import save_extraction_error
        save_extraction_error(
            url=url,
            error_code=error_code,
            error_message=error_message,
            http_status=http_status,
            raw_response_preview=preview,
        )
    except Exception:
        pass


@tool
def fetch_source(url: str) -> str:
    """Fetch and extract all useful content from a URL in one call.

    Automatically detects the content type and uses the right strategy:
    - HTML pages: extracts text, tables, and links to downloadable data files.
      If the page requires JavaScript, uses a headless browser automatically.
    - PDF files: extracts text from all pages.
    - CSV files: returns raw CSV content.
    - JSON/API responses: returns raw JSON.
    - Excel files: reads all sheets and rows.

    Errors (HTTP failures, timeouts, blocked access, etc.) are automatically
    logged for human review.

    Returns a combined result with text content, any tables found, and links
    to downloadable data files (JSON, CSV, Excel, PDF, etc.) on the page.
    """
    # Detect content type via HEAD
    try:
        head = httpx.head(url, timeout=15, follow_redirects=True, headers=HEADERS)
        content_type = head.headers.get("content-type", "").lower()
    except httpx.HTTPStatusError as exc:
        code, msg = _classify_http_error(exc)
        _log_error(url, code, msg, http_status=exc.response.status_code)
        return f"[ERROR] {msg}"
    except httpx.HTTPError as exc:
        code, msg = _classify_connection_error(exc)
        _log_error(url, code, msg)
        return f"[ERROR] {msg}"

    try:
        # PDF
        if "pdf" in content_type or url.lower().endswith(".pdf"):
            result = _extract_pdf(url)
            if "no extractable text" in result.lower():
                _log_error(url, "missing_data", "PDF contains no extractable text", preview=result)
            return result[:10000]

        # CSV
        if "csv" in content_type or url.lower().endswith(".csv"):
            resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
            resp.raise_for_status()
            return resp.text[:10000]

        # JSON
        if "json" in content_type or url.lower().endswith(".json"):
            resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
            resp.raise_for_status()
            return resp.text[:10000]

        # Excel
        if any(x in content_type for x in ["spreadsheet", "excel"]) or url.lower().endswith((".xlsx", ".xls")):
            result = _extract_xlsx(url)
            if "no data found" in result.lower():
                _log_error(url, "missing_data", "Excel file contains no data", preview=result)
            return result[:10000]

        # HTML — fetch once, extract everything
        html = _get_html(url)

        pw_error = ""
        if "<!--PLAYWRIGHT_ERROR:" in html:
            marker_end = html.index("-->")
            pw_error = html[len("<!--PLAYWRIGHT_ERROR:"):marker_end]
            html = html[marker_end + 3:]
            _log_error(url, "js_rendered", f"Playwright failed: {pw_error}", preview=html[:500])

        sections = []

        # Text content
        text = _extract_text_from_html(html)
        if text:
            # Check for anti-scraping / bot detection signals
            text_lower = text.lower()
            if any(sig in text_lower for sig in ["captcha", "verify you are human", "are you a robot"]):
                _log_error(url, "captcha", "CAPTCHA or bot detection detected", preview=text[:500])
            if any(sig in text_lower for sig in ["access denied", "blocked", "not allowed"]):
                _log_error(url, "ip_banned", "Access appears blocked", preview=text[:500])
            sections.append(text)

        # Tables
        tables = _extract_tables_from_html(html)
        if tables:
            sections.append(f"\n--- TABLES ---\n{tables}")

        # Data file links
        data_links = _extract_data_links(html, url)
        if data_links:
            sections.append(f"\n--- DATA LINKS ---\n{data_links}")

        if pw_error:
            sections.insert(0, f"[PLAYWRIGHT ERROR] Page requires JavaScript but Playwright failed: {pw_error}\n")

        result = "\n\n".join(sections)
        if not result:
            _log_error(url, "missing_data", "No content found on page")
            return "No content found on this page."

        return result[:10000]

    except httpx.HTTPStatusError as exc:
        code, msg = _classify_http_error(exc)
        _log_error(url, code, msg, http_status=exc.response.status_code)
        return f"[ERROR] {msg}"
    except httpx.HTTPError as exc:
        code, msg = _classify_connection_error(exc)
        _log_error(url, code, msg)
        return f"[ERROR] {msg}"
    except Exception as exc:
        _log_error(url, "missing_data", f"Unexpected error: {exc}")
        return f"[ERROR] Unexpected error: {exc}"


@tool
def check_url(url: str) -> str:
    """Check if a URL is accessible and returns a valid response.
    Use this to verify that a data source URL is live before attempting extraction.
    Returns status code and content type.
    """
    try:
        resp = httpx.head(url, timeout=15, follow_redirects=True, headers=HEADERS)
        content_type = resp.headers.get("content-type", "unknown")
        return f"Status: {resp.status_code}, Content-Type: {content_type}"
    except httpx.HTTPError as e:
        return f"Error: {e}"
