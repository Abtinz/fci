"""File download tools for various data formats."""

from __future__ import annotations

import io
import csv
import httpx
import openpyxl
from langchain_core.tools import tool

HEADERS = {
    "User-Agent": "Vision1M-Scorecard-Bot/1.0 (research; University of Waterloo)"
}


@tool
def download_xlsx(url: str, max_rows: int = 50) -> str:
    """Download an Excel (.xlsx) file and return its contents as text.
    Use this when the data source is an Excel spreadsheet.
    Returns sheet names and rows formatted as pipe-separated values.
    Args:
        url: Direct URL to the .xlsx file.
        max_rows: Maximum rows to read per sheet (default 50).
    """
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append(f"... truncated at {max_rows} rows")
                break
            row_str = " | ".join(str(c) if c is not None else "" for c in row)
            if row_str.strip(" |"):
                rows.append(row_str)
        if rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
    wb.close()

    return "\n\n".join(parts)[:10000] if parts else "No data found in spreadsheet."


@tool
def download_csv(url: str, max_rows: int = 100) -> str:
    """Download a CSV file and return its contents as text.
    Use this when the data source is a CSV file.
    Returns rows formatted as pipe-separated values.
    Args:
        url: Direct URL to the .csv file.
        max_rows: Maximum rows to read (default 100).
    """
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    reader = csv.reader(io.StringIO(resp.text))
    rows = []
    for i, row in enumerate(reader):
        if i >= max_rows:
            rows.append(f"... truncated at {max_rows} rows")
            break
        rows.append(" | ".join(row))

    return "\n".join(rows)[:10000] if rows else "No data found in CSV."


@tool
def download_file(url: str) -> str:
    """Download any file and return basic info about it.
    Use this to inspect a file's type and size before deciding how to parse it.
    Returns content type, size, and first 500 chars if text-based.
    """
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    content_type = resp.headers.get("content-type", "unknown")
    size = len(resp.content)

    info = f"Content-Type: {content_type}\nSize: {size} bytes\n"

    if "text" in content_type or "json" in content_type or "csv" in content_type:
        info += f"\nPreview:\n{resp.text[:500]}"

    return info
